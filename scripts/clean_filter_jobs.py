#!/usr/bin/env python3
"""Clean, filter, and summarize recruitment posting CSV files."""

from __future__ import annotations

import argparse
import csv
import re
from collections import Counter
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill


FIELD_ALIASES = {
    "title": ["title", "岗位", "职位", "职位名称", "招聘岗位", "工作岗位", "岗位名称", "job_title", "name"],
    "company": ["company", "公司", "公司名称", "招聘单位", "企业名称", "用人单位", "employer"],
    "city": ["city", "城市", "工作城市", "地点", "工作地点", "所在地", "地区", "location"],
    "salary": ["salary", "薪资", "薪酬", "工资", "薪资待遇", "月薪", "年薪", "待遇", "compensation"],
    "job_type": ["job_type", "类型", "岗位类型", "招聘类型", "职位类别", "工作性质", "用工形式"],
    "requirements": ["requirements", "要求", "任职要求", "描述", "岗位要求", "工作内容", "岗位职责", "招聘要求", "description"],
    "publish_time": ["publish_time", "发布时间", "发布日期", "更新时间", "更新日期", "date"],
    "detail_url": ["detail_url", "url", "链接", "详情链接", "原始URL", "职位链接", "申请链接", "详情页"],
    "source": ["source", "来源", "平台", "数据来源", "招聘平台"],
}


def pick(row: dict[str, str], field: str) -> str:
    lowered = {key.lower(): key for key in row.keys()}
    for alias in FIELD_ALIASES[field]:
        if alias in row:
            return row.get(alias, "").strip()
        key = lowered.get(alias.lower())
        if key:
            return row.get(key, "").strip()
    return ""


def parse_salary(text: str) -> tuple[float | None, float | None, str]:
    raw = (text or "").replace(" ", "")
    if not raw:
        return None, None, ""

    unit = "unknown"
    multiplier = 1.0
    if "万" in raw:
        unit = "monthly_yuan"
        multiplier = 10000.0
    elif "千" in raw or "k" in raw.lower():
        unit = "monthly_yuan"
        multiplier = 1000.0
    elif "元" in raw:
        unit = "yuan"

    numbers = [float(x) for x in re.findall(r"\d+(?:\.\d+)?", raw)]
    if not numbers:
        return None, None, unit
    if len(numbers) == 1:
        return numbers[0] * multiplier, numbers[0] * multiplier, unit
    return numbers[0] * multiplier, numbers[1] * multiplier, unit


def split_terms(value: str) -> list[str]:
    return [item.strip() for item in re.split(r"[,，;；|]", value or "") if item.strip()]


def clean_rows(rows: list[dict[str, str]]) -> tuple[list[dict[str, str]], int]:
    cleaned: list[dict[str, str]] = []
    seen: set[str] = set()
    duplicates = 0

    for row in rows:
        item = {field: pick(row, field) for field in FIELD_ALIASES}
        salary_min, salary_max, salary_unit = parse_salary(item["salary"])
        dedupe_key = "|".join(
            [
                item["detail_url"].lower(),
                item["title"].lower(),
                item["company"].lower(),
                item["city"].lower(),
            ]
        )
        if dedupe_key in seen:
            duplicates += 1
            continue
        seen.add(dedupe_key)
        item.update(
            {
                "salary_min": "" if salary_min is None else f"{salary_min:.0f}",
                "salary_max": "" if salary_max is None else f"{salary_max:.0f}",
                "salary_unit": salary_unit,
                "is_salary_known": "yes" if salary_min is not None else "no",
                "dedupe_key": dedupe_key,
            }
        )
        cleaned.append(item)
    return cleaned, duplicates


def filter_rows(
    rows: list[dict[str, str]],
    cities: list[str],
    keywords: list[str],
    salary_min: float | None,
) -> list[dict[str, str]]:
    selected = []
    for row in rows:
        haystack = " ".join([row.get("title", ""), row.get("requirements", ""), row.get("company", "")]).lower()
        matched = [term for term in keywords if term.lower() in haystack]
        city_ok = not cities or any(city in row.get("city", "") for city in cities)
        keyword_ok = not keywords or bool(matched)
        value = float(row["salary_min"]) if row.get("salary_min") else None
        salary_ok = salary_min is None or value is None or value >= salary_min
        if city_ok and keyword_ok and salary_ok:
            row = dict(row)
            row["matched_keywords"] = ",".join(matched)
            row["is_target_city"] = "yes" if city_ok else "no"
            selected.append(row)
    return selected



def write_xlsx(path: Path, rows: list[dict[str, str]]) -> None:
    fields = list(FIELD_ALIASES.keys()) + [
        "salary_min",
        "salary_max",
        "salary_unit",
        "matched_keywords",
        "is_target_city",
        "is_salary_known",
    ]
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()
    ws.title = "Jobs"

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    link_font = Font(color="0563C1", underline="single")

    for col_idx, field in enumerate(fields, 1):
        cell = ws.cell(row=1, column=col_idx, value=field)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, record in enumerate(rows, 2):
        for col_idx, field in enumerate(fields, 1):
            value = record.get(field, "")
            cell = ws.cell(row=row_idx, column=col_idx)
            if field == "detail_url" and value.startswith("http"):
                cell.value = value
                cell.font = link_font
                cell.hyperlink = value
            elif field in ("salary_min", "salary_max") and value:
                try:
                    cell.value = float(value)
                    cell.number_format = '#,##0'
                except (ValueError, TypeError):
                    cell.value = value
            else:
                cell.value = value

    for col_idx, field in enumerate(fields, 1):
        max_len = len(field)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2, max_row=len(rows) + 1):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 3, 60)

    wb.save(path)

def write_csv(path: Path, rows: list[dict[str, str]]) -> None:
    fields = list(FIELD_ALIASES.keys()) + [
        "salary_min",
        "salary_max",
        "salary_unit",
        "matched_keywords",
        "is_target_city",
        "is_salary_known",
        "dedupe_key",
    ]
    with path.open("w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.DictWriter(fh, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def write_report(path: Path, raw_count: int, duplicate_count: int, cleaned: list[dict[str, str]], filtered: list[dict[str, str]]) -> None:
    city_counts = Counter(row.get("city", "") or "unknown" for row in filtered)
    keyword_counts = Counter()
    for row in filtered:
        keyword_counts.update(split_terms(row.get("matched_keywords", "")))
    salary_known = sum(1 for row in filtered if row.get("is_salary_known") == "yes")

    lines = [
        "# Job Distribution Report",
        "",
        f"- Raw rows: {raw_count}",
        f"- Cleaned rows: {len(cleaned)}",
        f"- Duplicate rows removed: {duplicate_count}",
        f"- Filtered rows retained: {len(filtered)}",
        f"- Filtered rows with parsed salary: {salary_known}",
        "",
        "## Top Cities",
        "",
    ]
    lines.extend(f"- {city}: {count}" for city, count in city_counts.most_common(10))
    lines.extend(["", "## Keyword Matches", ""])
    if keyword_counts:
        lines.extend(f"- {keyword}: {count}" for keyword, count in keyword_counts.most_common(20))
    else:
        lines.append("- No keyword filter was applied or no keyword matches were found.")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("input_csv")
    parser.add_argument("--out-dir", default="outputs/jobs")
    parser.add_argument("--cities", default="")
    parser.add_argument("--keywords", default="")
    parser.add_argument("--salary-min", type=float, default=None)
    parser.add_argument("--xlsx", action="store_true", help="Export to .xlsx instead of .csv")
    args = parser.parse_args()

    input_path = Path(args.input_csv)
    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    with input_path.open("r", newline="", encoding="utf-8-sig") as fh:
        rows = list(csv.DictReader(fh))

    cleaned, duplicate_count = clean_rows(rows)
    filtered = filter_rows(cleaned, split_terms(args.cities), split_terms(args.keywords), args.salary_min)

    if args.xlsx:
        write_xlsx(out_dir / "cleaned_jobs.xlsx", cleaned)
        write_xlsx(out_dir / "filtered_jobs.xlsx", filtered)
    else:
        write_csv(out_dir / "cleaned_jobs.csv", cleaned)
        write_csv(out_dir / "filtered_jobs.csv", filtered)
    write_report(out_dir / "job_distribution.md", len(rows), duplicate_count, cleaned, filtered)


if __name__ == "__main__":
    main()
